
import imaplib
import email
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
import pytz
from email.utils import parsedate_to_datetime
from email.header import decode_header
from pathlib import Path
from config import EMAIL, PASSWORD, IMAP_SERVER
import chardet  # Importar a biblioteca chardet

# Função para decodificar o header do e-mail


def decode_email_header(header):
    decoded_header = decode_header(header)
    decoded_string = ""
    for part, encoding in decoded_header:
        if isinstance(part, bytes):
            if encoding:
                decoded_string += part.decode(encoding)
            else:
                decoded_string += part.decode()
        else:
            decoded_string += part
    return decoded_string

def decode_email_header(header):
    decoded_header = decode_header(header)
    decoded_string = ""
    for part, encoding in decoded_header:
        if isinstance(part, bytes):
            decoded_string += part.decode(encoding) if encoding else part.decode()
        else:
            decoded_string += part
    return decoded_string

def detect_encoding(data):
    """Detecta a codificação dos dados binários usando chardet."""
    result = chardet.detect(data)
    return result['encoding']

def get_email_body(msg):
    """Extrai o corpo do e-mail em texto simples ou HTML, lidando com diferentes codificações."""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = part.get("Content-Disposition", "")
            if content_type in ["text/plain", "text/html"] and "attachment" not in content_disposition:
                payload = part.get_payload(decode=True)
                # Detecção de codificação se não especificada
                charset = part.get_content_charset() or detect_encoding(payload)
                try:
                    return payload.decode(charset)
                except UnicodeDecodeError:
                    # Se ocorrer um erro na decodificação, tente com um fallback
                    return payload.decode(charset, errors='replace')
    else:
        # Para mensagens não multipart
        content_type = msg.get_content_type()
        payload = msg.get_payload(decode=True)
        charset = msg.get_content_charset() or detect_encoding(payload)
        try:
            return payload.decode(charset)
        except UnicodeDecodeError:
            return payload.decode(charset, errors='replace')


mail = imaplib.IMAP4_SSL(IMAP_SERVER)
mail.login(EMAIL, PASSWORD)
"""
# Listar as pastas/marcadores disponíveis
print("Listando pastas/marcadores disponíveis:")
folders = mail.list()
for folder in folders[1]:
    print(folder.decode())
"""
# Selecionar um marcador específico, por exemplo: 'Important'
mail.select('Grad/Aplicada')  # Atenção às aspas e ao caminho completo se necessário

# Selecionar a caixa de entrada
# mail.select('inbox')

# Imprimir a expressão de filtro usada
# print("Expressão de filtro: (SINCE 30 days ago)")

# Buscar pelos últimos 5 e-mails recebidos
# result, data = mail.search(None, 'ALL')

date_week_ago = (datetime.now() - timedelta(days=1500)).strftime('%d-%b-%Y')

# Criar o critério de busca para emails desde uma semana atrás
typ, data = mail.search(None, '(SINCE "{}")'.format(date_week_ago))

# Criar o critério de busca para emails desde uma semana atrás
#typ, data = mail.search(None, 'ALL')

email_ids = data[0].split() # Pegar os últimos e-mails da semana

# Criar uma nova planilha Excel
wb = Workbook()
ws = wb.active

# Adicionar cabeçalhos à planilha
ws.append(["Remetente", "Assunto", 'Texto',"Data", "ID", "Respondido"])

# Função para verificar se o email foi respondido
def check_if_replied(msg_id):
    result, data = mail.fetch(msg_id, "(FLAGS)")
    flags = data[0].decode('utf-8')
    if "\\Answered" in flags:
        return "Sim"
    else:
        return "Não"

# Carregar a planilha existente ou criar uma nova
excel_file = Path("Emails_Aplicada.xlsx")
if excel_file.exists():
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Remetente", "Assunto", "Texto","Data", "ID", "Respondido"])

# Ler os IDs existentes na planilha
existing_ids = {row[3] for row in ws.iter_rows(min_row=2, max_col=4, values_only=True)}

# Adicionar novos e-mails à planilha
for num in email_ids:
    print(f'Acesando email id {num}')
    msg_id = num.decode('utf-8')
    if msg_id not in existing_ids:
        result, data = mail.fetch(num, '(RFC822)')
        raw_email = data[0][1]
        msg = email.message_from_bytes(raw_email)
        sender = decode_email_header(msg['From'])
        subject = decode_email_header(msg['Subject'])
        body = subject = decode_email_header(msg['Subject'])
        raw_date = msg['Date']
        body = get_email_body(msg)
        date_obj = parsedate_to_datetime(raw_date)
        date_obj = date_obj.astimezone(pytz.timezone('GMT'))
        date_str = date_obj.strftime('%Y-%m-%d %H:%M:%S')
        replied = check_if_replied(msg_id)
        ws.append([sender, subject, body, date_str, msg_id, replied])

# Salvar a planilham
wb.save(excel_file)
print('\nDone!!!!')